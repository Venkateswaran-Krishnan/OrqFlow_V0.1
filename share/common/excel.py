from __future__ import annotations

import logging
import importlib.util
from pathlib import Path
from typing import Any
from zipfile import BadZipFile

import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException

try:
    from .exceptions import CommonUtilityError
except ImportError:  # Supports direct importlib loading of this file.
    exception_path = Path(__file__).with_name("exceptions.py")
    spec = importlib.util.spec_from_file_location("_common_exceptions", exception_path)
    if spec is None or spec.loader is None:
        raise
    exception_module = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(exception_module)
    CommonUtilityError = exception_module.CommonUtilityError


logger = logging.getLogger("common.excel")


def read_excel_dataframe(path: str | Path, sheet_name: str | int | None = None) -> pd.DataFrame:
    """Read an Excel worksheet as a pandas DataFrame without writing to the file."""
    excel_path = Path(path).expanduser()
    if not excel_path.exists():
        raise CommonUtilityError(f"Excel file not found: {excel_path}")
    if not excel_path.is_file():
        raise CommonUtilityError(f"Excel path is not a file: {excel_path}")

    workbook = None
    try:
        logger.debug("Reading Excel file as DataFrame: path=%s, sheet_name=%s", excel_path, sheet_name)
        workbook = load_workbook(excel_path, read_only=True, data_only=True)
        worksheet = _select_worksheet(workbook, sheet_name)
        rows = _trim_trailing_blank_rows(list(worksheet.iter_rows(values_only=True)))
        if not rows:
            raise CommonUtilityError(f"Excel sheet is empty: {worksheet.title}")

        headers = list(rows[0])
        if _is_blank_row(headers):
            raise CommonUtilityError(f"Excel sheet header row is empty: {worksheet.title}")

        data_rows = [list(row) for row in rows[1:]]
        return pd.DataFrame(data_rows, columns=headers)
    except CommonUtilityError:
        raise
    except PermissionError as error:
        raise CommonUtilityError(f"Excel file could not be opened due to permission or lock: {excel_path}") from error
    except (InvalidFileException, BadZipFile, OSError, ValueError) as error:
        raise CommonUtilityError(f"Excel file could not be read: {excel_path}. {error}") from error
    except Exception as error:
        logger.exception("Unexpected Excel utility failure")
        raise CommonUtilityError(f"Unexpected error while reading Excel file: {excel_path}. {error}") from error
    finally:
        if workbook is not None:
            workbook.close()


def _select_worksheet(workbook: Any, sheet_name: str | int | None) -> Any:
    if sheet_name is None:
        if not workbook.worksheets:
            raise CommonUtilityError("Excel workbook does not contain any worksheets")
        return workbook.worksheets[0]

    if isinstance(sheet_name, int):
        try:
            return workbook.worksheets[sheet_name]
        except IndexError as error:
            raise CommonUtilityError(f"Excel sheet index not found: {sheet_name}") from error

    if sheet_name not in workbook.sheetnames:
        raise CommonUtilityError(f"Excel sheet not found: {sheet_name}")
    return workbook[sheet_name]


def _trim_trailing_blank_rows(rows: list[tuple[Any, ...]]) -> list[tuple[Any, ...]]:
    while rows and _is_blank_row(rows[-1]):
        rows.pop()
    return rows


def _is_blank_row(row: list[Any] | tuple[Any, ...]) -> bool:
    return all(value is None or value == "" for value in row)
